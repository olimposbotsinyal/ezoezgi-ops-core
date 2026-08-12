# analytics/performance_report.py

import json
import os
import pandas as pd
import logging
from datetime import datetime, timezone, timedelta
from typing import Optional, List, Dict, Any
import io
import math
import re
import shutil

# Excel işlemleri için güvenli import
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, Series
except ImportError:
    openpyxl = None
    Font = PatternFill = Alignment = Border = Side = None
    get_column_letter = None
    BarChart = Reference = Series = None

logger = logging.getLogger(__name__)


class PerformanceReport:
    def __init__(self, signals_file: str = None):
        """
        Raporlama sınıfı. Hem ana dosyayı, hem arşivi hem de aktif sinyalleri yönetir.
        """
        self.base_dir = "analytics"
        self.archive_dir = "alarm_raporlari"  # Arşiv kök dizini

        # Kapanmış sinyaller
        self.main_file = signals_file or os.path.join(self.base_dir, "closed_signals_state.json")

        # Aktif sinyaller
        self.active_file = "active_signals_state.json"
        if not os.path.exists(self.active_file):
            self.active_file = os.path.join(self.base_dir, "active_signals_state.json")

        # Klasörleri oluştur
        os.makedirs(self.base_dir, exist_ok=True)
        os.makedirs(self.archive_dir, exist_ok=True)

        self.df = pd.DataFrame()
        # Varsayılan olarak ana dosyayı yükle
        self.load_data(self.main_file)

    @staticmethod
    def _safe_read_json(filepath: str) -> List[Dict]:
        """JSON dosyasını NaN/Infinity hatalarını temizleyerek okur."""
        if not os.path.exists(filepath):
            return []
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()
            # Regex ile temizlik
            content = re.sub(r':\s*NaN\b', ': null', content)
            content = re.sub(r':\s*Infinity\b', ': null', content)
            content = re.sub(r':\s*-Infinity\b', ': null', content)

            data = json.loads(content)
            if isinstance(data, list): return data
            if isinstance(data, dict):
                return data.get('closed_signals', data.get('signals', [data]))
            return []
        except Exception as e:
            logger.error(f"JSON okuma hatası ({filepath}): {e}")
            return []

    def load_data(self, filepath: str):
        """Belirtilen JSON dosyasından verileri yükler."""
        data = self._safe_read_json(filepath)
        if data:
            self.df = pd.DataFrame(data)
            if not self.df.empty:
                if 'signal_time' in self.df.columns:
                    self.df['signal_time'] = pd.to_datetime(self.df['signal_time'], errors='coerce', utc=True)
                if 'closed_time' in self.df.columns:
                    self.df['closed_time'] = pd.to_datetime(self.df['closed_time'], errors='coerce', utc=True)
        else:
            self.df = pd.DataFrame()

    def merge_active_signals(self):
        """Mevcut DataFrame'e aktif sinyalleri ekler."""
        active_data = self._safe_read_json(self.active_file)
        if not active_data: return

        try:
            df_active = pd.DataFrame(active_data)
            if 'signal_time' in df_active.columns:
                df_active['signal_time'] = pd.to_datetime(df_active['signal_time'], errors='coerce', utc=True)

            df_active['is_active_trade'] = True
            df_active['active'] = True

            if self.df.empty:
                self.df = df_active
            else:
                self.df = pd.concat([self.df, df_active], ignore_index=True)

            # Tekrarları önle
            if 'signal_id' in self.df.columns:
                if 'closed_time' in self.df.columns:
                    self.df['has_closed_time'] = self.df['closed_time'].notna()
                    self.df = self.df.sort_values(by='has_closed_time', ascending=False)
                    self.df = self.df.drop(columns=['has_closed_time'])
                self.df = self.df.drop_duplicates(subset=['signal_id'], keep='first')

        except Exception as e:
            logger.error(f"Aktif sinyalleri birleştirme hatası: {e}")

    def archive_old_signals(self):
        """24 saatten eski verileri arşivler ve 1 yıldan eskileri siler."""
        all_data = self._safe_read_json(self.main_file)
        if not all_data: return "Arşivlenecek veri yok."

        now = datetime.now(timezone.utc)
        cutoff_time = now - timedelta(hours=24)

        to_archive = []
        to_keep = []
        archived_count = 0

        for item in all_data:
            if item.get('is_active_trade', False) or item.get('active', False):
                to_keep.append(item)
                continue

            time_val = item.get('closed_time') or item.get('signal_time')
            if not time_val:
                to_keep.append(item)
                continue

            try:
                if isinstance(time_val, str):
                    dt_val = pd.to_datetime(time_val)
                    if dt_val.tzinfo is None: dt_val = dt_val.replace(tzinfo=timezone.utc)
                else: dt_val = time_val

                if dt_val < cutoff_time:
                    to_archive.append(item)
                else:
                    to_keep.append(item)
            except (ValueError, TypeError):
                to_keep.append(item)

        if to_archive:
            df_archive = pd.DataFrame(to_archive)
            df_archive['temp_time'] = pd.to_datetime(df_archive['signal_time'], errors='coerce', utc=True)
            grouped = df_archive.groupby([df_archive['temp_time'].dt.year, df_archive['temp_time'].dt.month])

            for (year, month), group in grouped:
                if pd.isna(year) or pd.isna(month): continue
                year, month = int(year), int(month)
                folder_path = os.path.join(self.archive_dir, str(year), f"{month:02d}")
                os.makedirs(folder_path, exist_ok=True)

                filename = f"closed_signals_{year}_{month:02d}.json"
                filepath = os.path.join(folder_path, filename)

                existing_archive = self._safe_read_json(filepath)
                existing_ids = {x.get('signal_id') for x in existing_archive if x.get('signal_id')}
                new_records = group.drop(columns=['temp_time']).to_dict('records')
                to_append = [x for x in new_records if x.get('signal_id') not in existing_ids]

                if to_append:
                    final_list = existing_archive + to_append
                    with open(filepath, 'w', encoding='utf-8') as f:
                        json.dump(final_list, f, indent=2, ensure_ascii=False, default=str)
                    archived_count += len(to_append)

            with open(self.main_file, 'w', encoding='utf-8') as f:
                json.dump(to_keep, f, indent=2, ensure_ascii=False, default=str)

        cleanup_cutoff_year = now.year - 1
        cleanup_cutoff_month = now.month
        if os.path.exists(self.archive_dir):
            for year_folder in os.listdir(self.archive_dir):
                if not year_folder.isdigit(): continue
                year_int = int(year_folder)
                year_path = os.path.join(self.archive_dir, year_folder)
                if year_int < cleanup_cutoff_year:
                    try: shutil.rmtree(year_path)
                    except OSError: pass
                elif year_int == cleanup_cutoff_year:
                    for month_folder in os.listdir(year_path):
                        if not month_folder.isdigit(): continue
                        month_int = int(month_folder)
                        if month_int < cleanup_cutoff_month:
                            try: shutil.rmtree(os.path.join(year_path, month_folder))
                            except OSError: pass

        self.df = pd.DataFrame(to_keep)
        return f"✅ {archived_count} işlem arşivlendi."

    def get_available_years(self) -> List[int]:
        """Veri olan yılları döndürür (Arşiv + Güncel)."""
        years = set()
        # Arşivden
        if os.path.exists(self.archive_dir):
            for item in os.listdir(self.archive_dir):
                if item.isdigit(): years.add(int(item))
        # Güncelden
        self.load_data(self.main_file)
        if not self.df.empty and 'signal_time' in self.df.columns:
            self.df['signal_time'] = pd.to_datetime(self.df['signal_time'], errors='coerce', utc=True)
            years.update(self.df['signal_time'].dt.year.dropna().astype(int).unique())
        return sorted(list(years), reverse=True)

    def get_available_months(self, year: int) -> List[int]:
        """Seçilen yıl için veri olan ayları döndürür (Arşiv + Güncel)."""
        months = set()
        # Arşivden
        year_path = os.path.join(self.archive_dir, str(year))
        if os.path.exists(year_path):
            for item in os.listdir(year_path):
                if item.isdigit():
                    m_path = os.path.join(year_path, item)
                    if any(f.endswith('.json') for f in os.listdir(m_path)):
                        months.add(int(item))
        # Güncelden
        self.load_data(self.main_file)
        if not self.df.empty and 'signal_time' in self.df.columns:
            self.df['signal_time'] = pd.to_datetime(self.df['signal_time'], errors='coerce', utc=True)
            current_year_data = self.df[self.df['signal_time'].dt.year == year]
            if not current_year_data.empty:
                months.update(current_year_data['signal_time'].dt.month.dropna().astype(int).unique())
        return sorted(list(months))

    def load_rolling_year(self):
        """Özel Aralık için 1 yıllık veri."""
        target_df = pd.DataFrame()
        now = datetime.now(timezone.utc)
        for i in range(13):
            date_check = now - timedelta(days=30 * i)
            year, month = date_check.year, date_check.month
            archive_path = os.path.join(self.archive_dir, str(year), f"{month:02d}",
                f"closed_signals_{year}_{month:02d}.json")
            if os.path.exists(archive_path):
                data = self._safe_read_json(archive_path)
                if data: target_df = pd.concat([target_df, pd.DataFrame(data)], ignore_index=True)

        self.load_data(self.main_file)
        self.merge_active_signals()
        if not self.df.empty: target_df = pd.concat([target_df, self.df], ignore_index=True)

        if not target_df.empty:
            if 'signal_id' in target_df.columns:
                target_df = target_df.drop_duplicates(subset=['signal_id'], keep='last')
            if 'signal_time' in target_df.columns:
                target_df['signal_time'] = pd.to_datetime(target_df['signal_time'], errors='coerce', utc=True)
                target_df = target_df[target_df['signal_time'] >= (now - timedelta(days=365))]
        self.df = target_df

    def load_specific_period(self, year: int, month: int, week: int = None):
        """Belirli dönemi yükler."""
        target_df = pd.DataFrame()
        archive_path = os.path.join(self.archive_dir, str(year), f"{month:02d}",
            f"closed_signals_{year}_{month:02d}.json")
        if os.path.exists(archive_path):
            data = self._safe_read_json(archive_path)
            if data: target_df = pd.DataFrame(data)

        current_active_df = pd.DataFrame(self._safe_read_json(self.main_file))
        if not current_active_df.empty and 'signal_time' in current_active_df.columns:
            current_active_df['signal_time'] = pd.to_datetime(current_active_df['signal_time'], errors='coerce',
                utc=True)
            mask = (current_active_df['signal_time'].dt.year == year) & (
                    current_active_df['signal_time'].dt.month == month)
            matches = current_active_df[mask]
            if not matches.empty: target_df = pd.concat([target_df, matches], ignore_index=True)

        if not target_df.empty:
            target_df['signal_time'] = pd.to_datetime(target_df['signal_time'], errors='coerce', utc=True)
            if 'signal_id' in target_df.columns:
                target_df = target_df.drop_duplicates(subset=['signal_id'], keep='last')
            if week is not None:
                start_day = (week - 1) * 7 + 1
                end_day = 31 if week == 4 else week * 7
                mask_week = (target_df['signal_time'].dt.day >= start_day) & (
                        target_df['signal_time'].dt.day <= end_day)
                target_df = target_df[mask_week]
        self.df = target_df

    @staticmethod
    def _calculate_active_pnl(row: pd.Series) -> float:
        try:
            targets, targets_hit = row.get('targets', []), row.get('targets_hit', [])
            entry_price = float(row.get('entry_price', 0) or 0)
            if entry_price == 0 or not targets: return 0.0

            last_hit_price = entry_price
            hit_any = False
            if isinstance(targets_hit, list):
                for i, is_hit in enumerate(targets_hit):
                    if is_hit and i < len(targets):
                        try:
                            last_hit_price = float(targets[i])
                            hit_any = True
                        except (ValueError, TypeError):
                            continue
            if not hit_any: return 0.0

            direction = row.get('signal_type', 'LONG')
            raw_gain = (last_hit_price - entry_price) / entry_price if direction == 'LONG' else (
                                                                                                        entry_price - last_hit_price) / entry_price
            lev = float(row.get('meta', {}).get('leverage', 10) or 10)
            return raw_gain * lev * 100
        except Exception:
            return 0.0

    def generate_detailed_telegram_table(self) -> str:
        """
        Telegram Tablosu (4 Kategori + Satır Boşluklu).
        """
        self.df = pd.DataFrame()
        self.load_data(self.main_file)
        self.merge_active_signals()

        if self.df.empty: return "⚠️ Görüntülenecek işlem bulunamadı."

        now = datetime.now(timezone.utc)
        cutoff_time = now - timedelta(hours=24)

        if 'closed_time' in self.df.columns:
            self.df['closed_time'] = pd.to_datetime(self.df['closed_time'], errors='coerce', utc=True)
        if 'signal_time' in self.df.columns:
            self.df['signal_time'] = pd.to_datetime(self.df['signal_time'], errors='coerce', utc=True)

        def is_relevant(row):
            if row.get('active') is True or row.get('is_active_trade') is True: return True
            c_time = row.get('closed_time')
            if pd.notnull(c_time) and c_time >= cutoff_time: return True
            s_time = row.get('signal_time')
            if pd.notnull(s_time) and s_time >= cutoff_time: return True
            return False

        mask = self.df.apply(is_relevant, axis=1)
        df = self.df[mask].copy()
        if df.empty: return "⚠️ Son 24 saatte aktif veya kapanan işlem bulunamadı."

        df = df.sort_values(by='signal_time', ascending=True)
        min_date = df['signal_time'].min().strftime('%d.%m')
        max_date = df['signal_time'].max().strftime('%d.%m')

        lines = [f"📊 <b>PERFORMANS RAPORU ({min_date} - {max_date})</b>"]
        lines.append("<code>NO  TARİH  SEMBOL   HEDEFLER     PNL</code>")
        lines.append("<code>" + "-" * 38 + "</code>")

        total_pnl = 0.0
        total_closed = 0

        # Kategoriler
        success_count = 0  # Tüm hedefler vuruldu
        semi_success_pos = 0  # En az 1 hedef + Kar
        semi_success_neg = 0  # En az 1 hedef + Zarar
        fail_count = 0  # Hiç hedef vurulmadı
        active_count = 0
        idx = 1

        for _, row in df.iterrows():
            is_active_flag = bool(row.get('active', False) or row.get('is_active_trade', False))
            stop_hit = bool(row.get('stop_loss_hit', False))
            if stop_hit or pd.notnull(row.get('closed_time')): is_active_flag = False

            sig_time = row.get('signal_time')
            date_str = sig_time.strftime('%H:%M') if pd.notnull(sig_time) else "--:--"
            symbol = str(row.get('symbol', 'UNK')).split('/')[0][:4]

            targets = row.get('targets', [])
            targets_hit = row.get('targets_hit', [])
            grid = ""

            # Hedef sayıları
            num_targets = len(targets) if isinstance(targets, list) else 3
            if num_targets == 0: num_targets = 3

            tp_hit_count = 0
            if isinstance(targets_hit, list):
                for i in range(num_targets):
                    is_hit = targets_hit[i] if i < len(targets_hit) else False
                    if is_hit:
                        grid += "✅"
                        tp_hit_count += 1
                    else:
                        grid += "⭕" if is_active_flag else "❌"
            else:
                grid = "⭕" * num_targets if is_active_flag else "❌" * num_targets

            lev = float(row.get('meta', {}).get('leverage', 10) or 10)
            pnl_str = ""
            result_icon = ""

            if is_active_flag:
                active_count += 1
                net_pct = self._calculate_active_pnl(row)
                result_icon = "⏳"
                pnl_str = f"%{int(net_pct):+} (A)"
            else:
                raw_pnl = float(row.get('realized_net_pct', 0) or 0)
                if math.isnan(raw_pnl): raw_pnl = 0.0
                net_pct = raw_pnl * lev
                total_pnl += net_pct
                total_closed += 1
                pnl_str = f"{int(net_pct):+}%"

                # --- YENİ KATEGORİZASYON MANTIĞI ---
                if tp_hit_count == num_targets and num_targets > 0:
                    # Tüm hedefler vuruldu -> Tam Başarılı
                    result_icon = "🏆"
                    success_count += 1
                elif 0 < tp_hit_count < num_targets:
                    # Arada bir yerde -> Yarı Başarılı
                    if net_pct > 0:
                        result_icon = "🟢"  # Pozitif Yarı Başarılı
                        semi_success_pos += 1
                    else:
                        result_icon = "🔴"  # Negatif Yarı Başarılı
                        semi_success_neg += 1
                else:
                    # Hiç hedef vurulmadı -> Başarısız
                    result_icon = "❌"
                    fail_count += 1

            line = f"<code>{idx:<2}) {date_str:<5} {symbol:<5} {grid} {result_icon} {pnl_str}</code>"
            lines.append(line)
            # OKUMA KOLAYLIĞI İÇİN BOŞ SATIR
            lines.append(" ")
            idx += 1

        lines.append("<code>" + "-" * 38 + "</code>")
        if active_count > 0: lines.append(f"🏃 <b>Devam Eden:</b> {active_count}")
        lines.append(f"🔒 <b>Kapatılan:</b> {total_closed}")

        total_semi = semi_success_pos + semi_success_neg

        lines.append(f"🏆 <b>Başarılı:</b> {success_count}")
        lines.append(f"🥇 <b>Yarı Başarılı:</b> {total_semi} ({semi_success_pos}+ / {semi_success_neg}-)")
        lines.append(f"🔴 <b>Başarısız:</b> {fail_count}")

        pnl_emoji = "🟢" if total_pnl > 0 else "🔴"
        lines.append(f"{pnl_emoji} <b>Net PnL:</b> %{total_pnl:.2f}")
        return "\n".join(lines)

    def filter_period(self, period: str = 'month') -> pd.DataFrame:
        if self.df.empty: return self.df
        now = datetime.now(timezone.utc)
        start_date = now
        if period == 'day':
            start_date = now - timedelta(hours=24)
        elif period == 'week':
            start_date = now - timedelta(days=7)
        elif period == 'month':
            start_date = now - timedelta(days=30)
        elif period == 'year':
            start_date = now - timedelta(days=365)
        elif period == 'all':
            return self.df.copy()
        mask = (self.df['signal_time'] >= start_date)
        return self.df[mask].copy()

    @staticmethod
    def summarize(df): return {}

    def generate_telegram_summary(self, days=15): return self.generate_detailed_telegram_table()

    @staticmethod
    def _calculate_duration(start, end):
        if pd.isna(start) or pd.isna(end): return "-"
        diff = end - start
        total_seconds = int(diff.total_seconds())
        days = total_seconds // 86400
        hours = (total_seconds % 86400) // 3600
        minutes = (total_seconds % 3600) // 60
        parts = []
        if days > 0: parts.append(f"{days}g")
        if hours > 0: parts.append(f"{hours}s")
        if minutes > 0: parts.append(f"{minutes}dk")
        if not parts: return "1dk<"
        return " ".join(parts)

    @staticmethod
    def cleanup_file(filepath: str):
        if filepath and os.path.exists(filepath):
            try:
                os.remove(filepath)
            except OSError:
                pass

    def export_excel_pro(self, start_date=None, end_date=None) -> Optional[str]:
        """
        Gelişmiş Excel Raporu (4 Kategori, Dashboard + Strateji Detayları).

        Güncellemeler (akış bozulmadan):
        - PnL doğrulaması: realized_effective_pct/realized_gross_pct/close_breakdown.gross_pct öncelikli
        - Kaldıraç: leverage_used (json) > meta.leverage > DB settings leverage > default
        - Margin: DB settings 'lot' (margin) > meta.entry_amount > default
        - Ek kolonlar: Borsa, Timeframe, Kar($), Zarar($)
        - Dashboard: Borsa özeti + grafik, Timeframe özeti + en iyi/en kötü 3 sembol
        """
        if openpyxl is None:
            logger.error("Excel oluşturma hatası: 'openpyxl' kütüphanesi yüklü değil.")
            return None

        df = self.df.copy()
        if df.empty:
            return None

        # Tarih formatlama
        if 'signal_time' in df.columns:
            df['signal_time'] = pd.to_datetime(df['signal_time'], errors='coerce', utc=True)
        if 'closed_time' in df.columns:
            df['closed_time'] = pd.to_datetime(df['closed_time'], errors='coerce', utc=True)

        # Tarih filtresi
        if start_date:
            if getattr(start_date, "tzinfo", None) is None:
                start_date = start_date.replace(tzinfo=timezone.utc)
            mask = (df['signal_time'] >= start_date)

            if end_date:
                if getattr(end_date, "tzinfo", None) is None:
                    end_date = end_date.replace(tzinfo=timezone.utc)
                mask = mask & (df['signal_time'] <= end_date)

            df = df[mask].copy()

        filename = f"Olimpos_Rapor_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        filepath = os.path.join(os.getcwd(), filename)

        rows_sheet1: List[Dict[str, Any]] = []
        rows_sheet2: List[Dict[str, Any]] = []

        default_leverage = 10.0
        default_margin_usdt = 100.0

        # Lazy import (circular riskine karşı)
        try:
            from data.olimpos_data import get_user_settings as _get_user_settings
        except ImportError:
            _get_user_settings = None

        # ---------------- HELPERS ----------------
        def _sf(val: Any, default: float = 0.0) -> float:
            """
            Safe-float: str/int/float/None için güvenli float dönüşümü.
            """
            try:
                if val is None:
                    return float(default)
                if isinstance(val, str):
                    s = val.strip()
                    if s in ("", "-", "none", "null"):
                        return float(default)
                    x = float(s)
                    return x if math.isfinite(x) else float(default)

                # int/float/bool vs
                x = float(val)
                return x if math.isfinite(x) else float(default)
            except (TypeError, ValueError):
                return float(default)

        settings_cache: Dict[tuple, Dict[str, Any]] = {}

        def _get_db_settings(user_id_val: Any, exchange_val: str) -> Dict[str, Any]:
            if not user_id_val:
                return {}

            ex = str(exchange_val or "").strip().lower()
            if not ex:
                return {}

            key = (int(user_id_val), ex)
            if key in settings_cache:
                return settings_cache[key]

            if _get_user_settings is None:
                settings_cache[key] = {}
                return {}

            try:
                # get_user_settings(user_id, exchange, username=None, channel_id=None, channel_name=None)
                s = _get_user_settings(int(user_id_val), ex)
                settings_cache[key] = s if isinstance(s, dict) else {}
                return settings_cache[key]
            except Exception:
                settings_cache[key] = {}
                return {}

        def _get_exchange(row_obj: pd.Series) -> str:
            meta_obj = row_obj.get('meta', {}) or {}
            meta_open_obj = row_obj.get('meta_at_open', {}) or {}

            ex_val = meta_obj.get("exchange") or meta_open_obj.get("exchange")
            if not ex_val:
                ex_val = (meta_obj.get("orderbook_cfg_ctx", {}) or {}).get("exchange")
            return str(ex_val) if ex_val else "-"

        def _get_timeframe(row_obj: pd.Series) -> str:
            tf_val = row_obj.get("timeframe")
            if tf_val:
                return str(tf_val)

            meta_obj = row_obj.get('meta', {}) or {}
            meta_open_obj = row_obj.get('meta_at_open', {}) or {}
            tf_val = meta_obj.get("timeframe") or meta_open_obj.get("timeframe")
            return str(tf_val) if tf_val else "-"

        def _get_source_display(row_obj: pd.Series) -> str:
            meta_obj = row_obj.get('meta', {}) or {}
            raw_source = meta_obj.get('source', row_obj.get('source', 'Bilinmiyor'))

            strat_id = meta_obj.get('strategy_id') or row_obj.get('strategy_id') or 'v1'
            strat_id = str(strat_id).upper()

            if isinstance(raw_source, str) and 'ai' in raw_source.lower():
                return f"🤖 AI - {strat_id}"
            if isinstance(raw_source, str) and 'strategy' in raw_source.lower():
                return f"📈 Strateji - {strat_id}"
            return f"{raw_source} - {strat_id}"

        def _pick_leverage(row_obj: pd.Series, settings_obj: Dict[str, Any]) -> float:
            # 1) json leverage_used (en doğru)
            lv = _sf(row_obj.get("leverage_used", 0), default=0.0)
            if lv > 0:
                return lv

            # 2) meta leverage
            meta_obj = row_obj.get('meta', {}) or {}
            lv = _sf(meta_obj.get("leverage", 0), default=0.0)
            if lv > 0:
                return lv

            # 3) db settings leverage (olası anahtarlar)
            for key_name in ("leverage", "kaldirac", "kaldirac_degeri", "kaldıraç"):
                if key_name in settings_obj:
                    lv = _sf(settings_obj.get(key_name), default=0.0)
                    if lv > 0:
                        return lv

            return float(default_leverage)

        def _pick_margin_usdt(row_obj: pd.Series, settings_obj: Dict[str, Any]) -> float:
            # 1) işlem bazlı kesin değer
            mv = _sf(row_obj.get("margin_used_usdt", 0), default=0.0)
            if mv > 0:
                return mv

            # 2) meta içinde saklanan işlem bazlı değer
            meta_obj = row_obj.get('meta', {}) or {}
            for k in ("entry_amount", "margin_usdt", "margin_used_usdt"):
                mv = _sf(meta_obj.get(k, 0), default=0.0)
                if mv > 0:
                    return mv

            # 3) DB fallback (tarihsel değil, sadece yoksa)
            mv = _sf(settings_obj.get("lot", 0), default=0.0)
            if mv > 0:
                return mv

            return float(default_margin_usdt)

        def _last_hit_target_price(target_list: Any, hit_list: Any) -> Optional[float]:
            if not isinstance(target_list, list) or not isinstance(hit_list, list):
                return None
            last_idx: Optional[int] = None
            for idx, hit_flag in enumerate(hit_list):
                if hit_flag is True and idx < len(target_list):
                    last_idx = idx
            if last_idx is None:
                return None
            price_val = _sf(target_list[last_idx], default=0.0)
            return price_val if price_val > 0 else None

        def _infer_exit_price(row_obj: pd.Series, entry_price_val: float, target_list: List[Any],
                hit_list: List[Any]) -> float:
            # varsa direkt exit_price
            ep = _sf(row_obj.get("exit_price", 0), default=0.0)
            if ep > 0:
                return ep

            exit_type_val = str(row_obj.get("exit_type") or "").upper()
            stop_hit_flag = bool(row_obj.get("stop_loss_hit", False))

            # STOP
            if stop_hit_flag or exit_type_val.startswith("STOP"):
                sl = _sf(row_obj.get("stop_loss", 0), default=0.0)
                return sl if sl > 0 else 0.0

            # TARGET / PARTIAL
            if "TARGET" in exit_type_val or "PART" in exit_type_val:
                last_tp = _last_hit_target_price(target_list, hit_list)
                if last_tp is not None:
                    return float(last_tp)

            # fallback
            last_tp = _last_hit_target_price(target_list, hit_list)
            if last_tp is not None:
                return float(last_tp)

            # en son: entry ile aynı kabul etmek yerine 0 bırak (Excel’de '-' basılacak)
            _ = entry_price_val  # entry kullanmıyoruz ama imzayı korumak istersen burada dursun
            return 0.0

        def _pick_base_realized_pct(row_obj: pd.Series) -> float:
            # Öncelik: effective > gross > close_breakdown.gross_pct > realized_net_pct
            for key_name in ("realized_effective_pct", "realized_gross_pct"):
                x = _sf(row_obj.get(key_name), default=0.0)
                if x != 0.0:
                    return x

            cb_obj = row_obj.get("close_breakdown", {}) or {}
            x = _sf(cb_obj.get("gross_pct"), default=0.0)
            if x != 0.0:
                return x

            return _sf(row_obj.get("realized_net_pct"), default=0.0)

        def _compute_raw_pct_from_prices(entry_val: float, exit_val: float, side: str) -> float:
            if entry_val <= 0 or exit_val <= 0:
                return 0.0
            if str(side).upper() == "LONG":
                return (exit_val - entry_val) / entry_val * 100.0
            return (entry_val - exit_val) / entry_val * 100.0

        def _ensure_numeric_col(frame: pd.DataFrame, col_name: str) -> None:
            if col_name not in frame.columns:
                frame[col_name] = 0.0
            frame[col_name] = pd.to_numeric(frame[col_name], errors="coerce").fillna(0.0)

        # ---------------- ROW LOOP ----------------
        for _, row in df.iterrows():
            meta_obj = row.get('meta', {}) or {}
            user_id_val = row.get("user_id")
            exchange = _get_exchange(row)
            settings_obj = _get_db_settings(user_id_val, exchange)

            signal_id = row.get('signal_id', '-')
            symbol = str(row.get('symbol', 'Unknown')).split(':')[0]
            direction = str(row.get('signal_type', 'LONG')).upper()

            source_display = _get_source_display(row)
            timeframe = _get_timeframe(row)

            open_time = row.get('signal_time')
            close_time = row.get('closed_time')
            duration = self._calculate_duration(open_time, close_time)

            entry_price = _sf(row.get('entry_price', 0), default=0.0)

            target_list = row.get('targets', []) or []
            hit_list = row.get('targets_hit', []) or []
            hit_times = row.get('targets_hit_times', []) or []

            leverage = _pick_leverage(row, settings_obj)
            margin_usdt = _pick_margin_usdt(row, settings_obj)

            entry_lot = _sf(meta_obj.get('entry_lot', 0), default=0.0)  # korunuyor

            exit_price = _infer_exit_price(row, entry_price, target_list, hit_list)

            num_targets = len(target_list) if isinstance(target_list, list) else 3
            if num_targets == 0:
                num_targets = 3

            tp_hit_count = sum(1 for v in hit_list if v is True) if isinstance(hit_list, list) else 0

            # TP datası
            tp_data: Dict[str, Any] = {}
            for tp_i in range(5):
                tp_idx = tp_i + 1
                tp_price = target_list[tp_i] if tp_i < len(target_list) else None
                tp_hit = bool(hit_list[tp_i]) if tp_i < len(hit_list) else False
                tp_hit_time = hit_times[tp_i] if tp_i < len(hit_times) else None

                status_icon = "✅" if tp_hit else "❌"

                tp_duration = "-"
                if tp_hit and tp_hit_time and pd.notnull(open_time):
                    try:
                        hit_dt = pd.to_datetime(tp_hit_time, utc=True, errors="coerce")
                        if pd.notnull(hit_dt):
                            tp_duration = self._calculate_duration(open_time, hit_dt)
                    except (TypeError, ValueError):
                        pass

                tp_pnl_str = "-"
                if tp_hit and tp_price and entry_price > 0:
                    tp_p = _sf(tp_price, default=0.0)
                    if tp_p > 0:
                        raw_move = (tp_p - entry_price) / entry_price if direction == "LONG" else (
                                                                                                              entry_price - tp_p) / entry_price
                        tp_pnl_str = f"%{raw_move * leverage * 100.0:.2f}"

                tp_data[f'tp{tp_idx}_price'] = tp_price
                tp_data[f'tp{tp_idx}_status'] = status_icon
                tp_data[f'tp{tp_idx}_dur'] = tp_duration
                tp_data[f'tp{tp_idx}_pnl'] = tp_pnl_str

            sl_price = _sf(row.get('stop_loss', 0), default=0.0)

            # PnL doğrulama
            raw_base_pct = _pick_base_realized_pct(row)  # kaldıraçsız (%)
            if raw_base_pct == 0.0:
                raw_base_pct = _compute_raw_pct_from_prices(entry_price, exit_price, direction)

            net_pnl_leveraged = raw_base_pct * leverage
            real_pnl_usdt = margin_usdt * (net_pnl_leveraged / 100.0)

            profit_usdt = real_pnl_usdt if real_pnl_usdt > 0 else 0.0
            loss_usdt = real_pnl_usdt if real_pnl_usdt < 0 else 0.0  # negatif

            # Sonuç kategorisi (mevcut mantık)
            if tp_hit_count == num_targets and num_targets > 0:
                result_text = "🏆 BAŞARILI (Tümü)"
            elif 0 < tp_hit_count < num_targets:
                result_text = "🟢 YARI BAŞARILI (+)" if net_pnl_leveraged > 0 else "🔴 YARI BAŞARILI (-)"
            else:
                result_text = "❌ BAŞARISIZ"

            date_display = open_time.strftime('%d.%m.%Y %H:%M') if pd.notnull(open_time) else '-'

            # Sheet1
            kaldirac_display = f"x{int(leverage)}" if float(leverage).is_integer() else f"x{leverage:g}"
            margin_display = f"${margin_usdt:.2f}".rstrip('0').rstrip('.')

            rows_sheet1.append({
                'Tarih':date_display,
                'Sembol':symbol,
                'Borsa':exchange,
                'Timeframe':timeframe,
                'Yön':direction,
                'Strateji Kaynağı':source_display,

                'Kaldıraç':kaldirac_display,
                'Marjin (USDT)':margin_display,

                'Lot':f"{entry_lot:.2f}" if entry_lot > 0 else "-",

                'Giriş':entry_price,
                'Çıkış':exit_price if exit_price > 0 else '-',
                'Süre':duration,

                'TP1 Fiyat':tp_data['tp1_price'], 'TP1 Durum':tp_data['tp1_status'], 'TP1 Süre':tp_data['tp1_dur'],
                'TP1 PnL':tp_data['tp1_pnl'],
                'TP2 Fiyat':tp_data['tp2_price'], 'TP2 Durum':tp_data['tp2_status'], 'TP2 Süre':tp_data['tp2_dur'],
                'TP2 PnL':tp_data['tp2_pnl'],
                'TP3 Fiyat':tp_data['tp3_price'], 'TP3 Durum':tp_data['tp3_status'], 'TP3 Süre':tp_data['tp3_dur'],
                'TP3 PnL':tp_data['tp3_pnl'],
                'TP4 Fiyat':tp_data['tp4_price'], 'TP4 Durum':tp_data['tp4_status'], 'TP4 Süre':tp_data['tp4_dur'],
                'TP4 PnL':tp_data['tp4_pnl'],
                'TP5 Fiyat':tp_data['tp5_price'], 'TP5 Durum':tp_data['tp5_status'], 'TP5 Süre':tp_data['tp5_dur'],
                'TP5 PnL':tp_data['tp5_pnl'],

                'Stop Loss':sl_price,
                'Sonuç':result_text,

                'Net PnL (%)':net_pnl_leveraged,
                'Gerçek PnL ($)':real_pnl_usdt,

                'Kar ($)':profit_usdt,
                'Zarar ($)':loss_usdt,

                'ID':signal_id
            })

            # Sheet2
            rows_sheet2.append({
                'ID':signal_id,
                'Sembol':symbol,
                'Borsa':exchange,
                'Timeframe':timeframe,
                'Strateji Kaynağı':source_display,
                'Teknik Skor':meta_obj.get('technical_score', 0),
                'AI Güveni':meta_obj.get('ai_confidence', 0),
                'RSI':meta_obj.get('rsi', 0),
                'Momentum':meta_obj.get('momentum', 0),
                'Sonuç':result_text,
                'Net PnL (%)':net_pnl_leveraged
            })

        df_sheet1 = pd.DataFrame(rows_sheet1)
        df_sheet2 = pd.DataFrame(rows_sheet2)

        # ---------------- WRITE EXCEL ----------------
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            if not df_sheet1.empty:
                df_sheet1.to_excel(writer, sheet_name='Detaylı İşlem Listesi', index=False)
            if not df_sheet2.empty:
                df_sheet2.to_excel(writer, sheet_name='Teknik Veriler', index=False)

            wb = writer.book
            ws_dash = wb.create_sheet('Genel Değerlendirme')

            # İstatistikler (mevcut)
            count_success = count_semi_pos = count_semi_neg = count_fail = 0
            net_total_pnl = 0.0
            real_pnl_total = 0.0
            best_symbols = pd.Series(dtype=float)
            worst_symbols = pd.Series(dtype=float)

            total_trades = len(df_sheet1)
            if total_trades > 0:
                success_mask = df_sheet1['Sonuç'].astype(str).str.contains(r'BAŞARILI \(Tümü\)', na=False)
                semi_pos_mask = df_sheet1['Sonuç'].astype(str).str.contains(r'YARI BAŞARILI \(\+\)', na=False)
                semi_neg_mask = df_sheet1['Sonuç'].astype(str).str.contains(r'YARI BAŞARILI \(\-\)', na=False)
                fail_mask = df_sheet1['Sonuç'].astype(str).str.contains(r'BAŞARISIZ', na=False)

                count_success = int(success_mask.sum())
                count_semi_pos = int(semi_pos_mask.sum())
                count_semi_neg = int(semi_neg_mask.sum())
                count_fail = int(fail_mask.sum())

                net_total_pnl = float(pd.to_numeric(df_sheet1['Net PnL (%)'], errors='coerce').fillna(0.0).sum())
                real_pnl_total = float(pd.to_numeric(df_sheet1['Gerçek PnL ($)'], errors='coerce').fillna(0.0).sum())

                symbol_grp = df_sheet1.groupby('Sembol')['Net PnL (%)'].sum().sort_values(ascending=False)
                best_symbols = symbol_grp.head(5)
                worst_symbols = symbol_grp.tail(5)

            header_font = Font(bold=True, size=14, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

            ws_dash['B2'] = "📊 GENEL İŞLEM ÖZETİ"
            ws_dash['B2'].font = header_font
            ws_dash['B2'].fill = header_fill

            summary_data = [
                ('Toplam İşlem', total_trades),
                ('Tam Başarılı (Tüm Hedefler)', count_success),
                ('Yarı Başarılı (Pozitif)', count_semi_pos),
                ('Yarı Başarılı (Negatif)', count_semi_neg),
                ('Başarısız (0 Hedef)', count_fail),
                ('Genel Başarı Oranı',
                    f"%{((count_success + count_semi_pos) / total_trades * 100):.1f}" if total_trades > 0 else "0")
            ]

            row_idx = 3
            for key_label, val in summary_data:
                ws_dash[f'B{row_idx}'] = key_label
                ws_dash[f'C{row_idx}'] = val
                row_idx += 1

            row_idx += 1
            ws_dash[f'B{row_idx}'] = "💰 FİNANSAL ÖZET"
            ws_dash[f'B{row_idx}'].font = header_font
            ws_dash[f'B{row_idx}'].fill = header_fill
            row_idx += 1

            fin_data = [
                ('Toplam PnL (%)', f"%{net_total_pnl:.2f}"),
                ('Ortalama PnL', f"%{(net_total_pnl / total_trades):.2f}" if total_trades > 0 else "0"),
                ('TOPLAM KASA KARI', f"${real_pnl_total:.2f}")
            ]
            for key_label, val in fin_data:
                ws_dash[f'B{row_idx}'] = key_label
                ws_dash[f'C{row_idx}'] = val
                if "TOPLAM" in key_label:
                    ws_dash[f'C{row_idx}'].font = Font(bold=True, size=12,
                        color="006100" if real_pnl_total > 0 else "9C0006")
                row_idx += 1

            ws_dash['E2'] = "🏆 EN İYİLER"
            ws_dash['E2'].font = header_font
            ws_dash['E2'].fill = PatternFill(start_color="9BBB59", fill_type="solid")
            rb = 3
            for sym, val in best_symbols.items():
                ws_dash[f'E{rb}'] = sym
                ws_dash[f'F{rb}'] = f"%{val:.2f}"
                rb += 1

            ws_dash['E10'] = "💀 EN KÖTÜLER"
            ws_dash['E10'].font = header_font
            ws_dash['E10'].fill = PatternFill(start_color="C0504D", fill_type="solid")
            rw = 11
            for sym, val in worst_symbols.items():
                ws_dash[f'E{rw}'] = sym
                ws_dash[f'F{rw}'] = f"%{val:.2f}"
                rw += 1

            # Strateji performans tablosu (mevcut)
            row_idx += 2
            ws_dash[f'B{row_idx}'] = "📈 STRATEJİ PERFORMANSI"
            ws_dash[f'B{row_idx}'].font = header_font
            ws_dash[f'B{row_idx}'].fill = PatternFill(start_color="808080", fill_type="solid")
            row_idx += 1

            chart_data = [["Strateji", "Başarılı", "Yarı(+)", "Yarı(-)", "Başarısız"]]
            strat_groups = None

            if not df_sheet1.empty:
                strat_groups = df_sheet1.groupby('Strateji Kaynağı')
                for s_name, grp in strat_groups:
                    c_s = len(grp[grp['Sonuç'].astype(str).str.contains(r'BAŞARILI \(Tümü\)', na=False)])
                    c_sp = len(grp[grp['Sonuç'].astype(str).str.contains(r'YARI BAŞARILI \(\+\)', na=False)])
                    c_sn = len(grp[grp['Sonuç'].astype(str).str.contains(r'YARI BAŞARILI \(\-\)', na=False)])
                    c_f = len(grp[grp['Sonuç'].astype(str).str.contains(r'BAŞARISIZ', na=False)])
                    chart_data.append([s_name, c_s, c_sp, c_sn, c_f])

                    ws_dash[f'B{row_idx}'] = s_name
                    ws_dash[f'C{row_idx}'] = f"B:{c_s} / Y+:{c_sp} / Y-:{c_sn} / F:{c_f}"
                    row_idx += 1

            if strat_groups is not None:
                row_idx += 2
                ws_dash[f'B{row_idx}'] = "⚡ STRATEJİ DETAY (EN İYİ/KÖTÜ)"
                ws_dash[f'B{row_idx}'].font = header_font
                ws_dash[f'B{row_idx}'].fill = PatternFill(start_color="60497A", fill_type="solid")
                row_idx += 1

                for s_name, grp in strat_groups:
                    grp_sorted = grp.sort_values(by='Net PnL (%)', ascending=False)
                    best_2 = grp_sorted.head(2)
                    worst_2 = grp_sorted.tail(2)

                    ws_dash[f'B{row_idx}'] = s_name
                    ws_dash[f'B{row_idx}'].font = Font(bold=True, underline="single")
                    row_idx += 1

                    ws_dash[f'B{row_idx}'] = "En İyiler:"
                    col_offset = 2
                    for _, row_best in best_2.iterrows():
                        c = ws_dash.cell(row=row_idx, column=1 + col_offset)
                        c.value = f"{row_best['Sembol']} (%{row_best['Net PnL (%)']:.0f})"
                        c.font = Font(color="006100")
                        col_offset += 1
                    row_idx += 1

                    ws_dash[f'B{row_idx}'] = "En Kötüler:"
                    col_offset = 2
                    for _, row_worst in worst_2.iterrows():
                        c = ws_dash.cell(row=row_idx, column=1 + col_offset)
                        c.value = f"{row_worst['Sembol']} (%{row_worst['Net PnL (%)']:.0f})"
                        c.font = Font(color="9C0006")
                        col_offset += 1
                    row_idx += 2

            # -------- BORSA ÖZETİ + GRAFİK --------
            row_idx += 1
            ws_dash[f'B{row_idx}'] = "🏦 BORSA ÖZETİ"
            ws_dash[f'B{row_idx}'].font = header_font
            ws_dash[f'B{row_idx}'].fill = PatternFill(start_color="1F4E79", fill_type="solid")
            row_idx += 1

            if not df_sheet1.empty and 'Borsa' in df_sheet1.columns:
                tmp_df = df_sheet1.copy()
                _ensure_numeric_col(tmp_df, 'Gerçek PnL ($)')
                _ensure_numeric_col(tmp_df, 'Kar ($)')
                _ensure_numeric_col(tmp_df, 'Zarar ($)')

                grp_ex = tmp_df.groupby('Borsa', dropna=False).agg(
                    islem_sayisi=('ID', 'count'),
                    kar_usd=('Kar ($)', 'sum'),
                    zarar_usd=('Zarar ($)', 'sum'),
                    net_usd=('Gerçek PnL ($)', 'sum')
                ).reset_index()

                grp_ex['zarar_abs'] = grp_ex['zarar_usd'].abs()
                exch_summary_df = grp_ex.sort_values(by='net_usd', ascending=False)

                ws_dash[f'B{row_idx}'] = "Borsa"
                ws_dash[f'C{row_idx}'] = "İşlem"
                ws_dash[f'D{row_idx}'] = "Kar($)"
                ws_dash[f'E{row_idx}'] = "Zarar($)"
                ws_dash[f'F{row_idx}'] = "Net($)"
                for col_letter in "BCDEF":
                    ws_dash[f'{col_letter}{row_idx}'].font = Font(bold=True)
                row_idx += 1

                # tablo
                for _, gx_row in exch_summary_df.iterrows():
                    ws_dash[f'B{row_idx}'] = str(gx_row['Borsa'])
                    ws_dash[f'C{row_idx}'] = int(gx_row['islem_sayisi'])
                    ws_dash[f'D{row_idx}'] = float(gx_row['kar_usd'])
                    ws_dash[f'E{row_idx}'] = float(gx_row['zarar_abs'])
                    ws_dash[f'F{row_idx}'] = float(gx_row['net_usd'])
                    row_idx += 1

                # grafik veri alanı (H sütunu)
                base_col = 8  # H
                start_r = 2
                ws_dash.cell(row=start_r, column=base_col, value="Borsa")
                ws_dash.cell(row=start_r, column=base_col + 1, value="Kar($)")
                ws_dash.cell(row=start_r, column=base_col + 2, value="Zarar($)")

                for ridx, gx_row in enumerate(exch_summary_df.itertuples(index=False), start=1):
                    ws_dash.cell(row=start_r + ridx, column=base_col, value=str(getattr(gx_row, "Borsa")))
                    ws_dash.cell(row=start_r + ridx, column=base_col + 1, value=float(getattr(gx_row, "kar_usd")))
                    ws_dash.cell(row=start_r + ridx, column=base_col + 2, value=float(getattr(gx_row, "zarar_abs")))

                # stacked bar
                try:
                    chart_ex = BarChart()
                    chart_ex.type = "col"
                    chart_ex.grouping = "stacked"
                    chart_ex.overlap = 100
                    chart_ex.title = "Borsa Bazlı Kar/Zarar ($)"
                    chart_ex.height = 10
                    chart_ex.width = 18

                    data_ref = Reference(
                        ws_dash,
                        min_col=base_col + 1,
                        min_row=start_r,
                        max_col=base_col + 2,
                        max_row=start_r + len(exch_summary_df)
                    )
                    cats_ref = Reference(
                        ws_dash,
                        min_col=base_col,
                        min_row=start_r + 1,
                        max_row=start_r + len(exch_summary_df)
                    )
                    chart_ex.add_data(data_ref, titles_from_data=True)
                    chart_ex.set_categories(cats_ref)
                    ws_dash.add_chart(chart_ex, "H20")
                except Exception:
                    pass

            # -------- TIMEFRAME ÖZETİ + EN İYİ/EN KÖTÜ 3 --------
            row_idx += 1
            ws_dash[f'B{row_idx}'] = "⏱️ TIMEFRAME ÖZETİ"
            ws_dash[f'B{row_idx}'].font = header_font
            ws_dash[f'B{row_idx}'].fill = PatternFill(start_color="375623", fill_type="solid")
            row_idx += 1

            if not df_sheet1.empty and 'Timeframe' in df_sheet1.columns:
                tf_tmp_df = df_sheet1.copy()
                _ensure_numeric_col(tf_tmp_df, 'Gerçek PnL ($)')
                _ensure_numeric_col(tf_tmp_df, 'Kar ($)')
                _ensure_numeric_col(tf_tmp_df, 'Zarar ($)')
                tf_tmp_df['ZararAbs ($)'] = tf_tmp_df['Zarar ($)'].abs()

                tf_grp_df = tf_tmp_df.groupby('Timeframe', dropna=False).agg(
                    islem_sayisi=('ID', 'count'),
                    kar_usd=('Kar ($)', 'sum'),
                    zarar_abs=('ZararAbs ($)', 'sum'),
                    net_usd=('Gerçek PnL ($)', 'sum')
                ).reset_index().sort_values(by='net_usd', ascending=False)

                ws_dash[f'B{row_idx}'] = "TF"
                ws_dash[f'C{row_idx}'] = "İşlem"
                ws_dash[f'D{row_idx}'] = "Kar($)"
                ws_dash[f'E{row_idx}'] = "Zarar($)"
                ws_dash[f'F{row_idx}'] = "Net($)"
                for col_letter in "BCDEF":
                    ws_dash[f'{col_letter}{row_idx}'].font = Font(bold=True)
                row_idx += 1

                for _, tf_row in tf_grp_df.iterrows():
                    ws_dash[f'B{row_idx}'] = str(tf_row['Timeframe'])
                    ws_dash[f'C{row_idx}'] = int(tf_row['islem_sayisi'])
                    ws_dash[f'D{row_idx}'] = float(tf_row['kar_usd'])
                    ws_dash[f'E{row_idx}'] = float(tf_row['zarar_abs'])
                    ws_dash[f'F{row_idx}'] = float(tf_row['net_usd'])
                    row_idx += 1

                row_idx += 1
                ws_dash[f'B{row_idx}'] = "🏅 TIMEFRAME EN İYİ / EN KÖTÜ 3 SEMBOL"
                ws_dash[f'B{row_idx}'].font = Font(bold=True)
                row_idx += 1

                tf_symbol_df = df_sheet1.copy()
                tf_symbol_df['Net PnL (%)'] = pd.to_numeric(tf_symbol_df['Net PnL (%)'], errors='coerce').fillna(0.0)

                for tf_name, tf_group in tf_symbol_df.groupby('Timeframe'):
                    ws_dash[f'B{row_idx}'] = f"TF: {tf_name}"
                    ws_dash[f'B{row_idx}'].font = Font(bold=True, underline="single")
                    row_idx += 1

                    sym_sum = tf_group.groupby('Sembol')['Net PnL (%)'].sum().sort_values(ascending=False)
                    best3 = sym_sum.head(3)
                    worst3 = sym_sum.tail(3)

                    ws_dash[f'B{row_idx}'] = "En İyiler:"
                    ccol = 3
                    for sym, val in best3.items():
                        c = ws_dash.cell(row=row_idx, column=ccol)
                        c.value = f"{sym} (%{val:.2f})"
                        c.font = Font(color="006100")
                        ccol += 1
                    row_idx += 1

                    ws_dash[f'B{row_idx}'] = "En Kötüler:"
                    ccol = 3
                    for sym, val in worst3.items():
                        c = ws_dash.cell(row=row_idx, column=ccol)
                        c.value = f"{sym} (%{val:.2f})"
                        c.font = Font(color="9C0006")
                        ccol += 1
                    row_idx += 2

            # Strateji grafiği (mevcut)
            start_r_chart = 2
            for ri, row_d in enumerate(chart_data):
                for cj, v in enumerate(row_d):
                    ws_dash.cell(row=start_r_chart + ri, column=16 + cj, value=v)

            try:
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.grouping = "stacked"
                chart.overlap = 100
                chart.title = "Strateji Performans Dağılımı"
                chart.height = 10
                chart.width = 15

                data_ref = Reference(ws_dash, min_col=17, min_row=2, max_row=1 + len(chart_data), max_col=20)
                cats_ref = Reference(ws_dash, min_col=16, min_row=3, max_row=1 + len(chart_data))

                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                ws_dash.add_chart(chart, "H2")
            except Exception:
                pass

            # Dashboard sütun genişliği
            ws_dash.column_dimensions['B'].width = 32
            ws_dash.column_dimensions['C'].width = 18
            ws_dash.column_dimensions['D'].width = 16
            ws_dash.column_dimensions['E'].width = 16
            ws_dash.column_dimensions['F'].width = 16
            ws_dash.column_dimensions['H'].width = 18

            # Renklendirme (mevcut mantık)
            def colorize_sheet(worksheet):
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell_obj in column:
                        try:
                            max_length = max(max_length, len(str(cell_obj.value)))
                        except Exception:
                            pass
                    worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

                green_font = Font(color="00B050", bold=True)
                red_font = Font(color="FF0000", bold=True)

                fin_green_font = Font(color="006100")
                fin_red_font = Font(color="9C0006")

                for excel_row in worksheet.iter_rows(min_row=2):
                    for cell_obj in excel_row:
                        val_str = str(cell_obj.value)

                        if "✅" in val_str or "🟢" in val_str or "🏆" in val_str:
                            cell_obj.font = green_font
                        elif "❌" in val_str or "🛑" in val_str or "🔴" in val_str:
                            cell_obj.font = red_font

                        if isinstance(cell_obj.value, (int, float)):
                            header_cell = worksheet.cell(row=1, column=cell_obj.column)
                            header_txt = str(header_cell.value) if header_cell.value else ""
                            if any(x in header_txt for x in ("PnL", "Net", "Kar", "Zarar")):
                                if cell_obj.value > 0:
                                    cell_obj.font = fin_green_font
                                elif cell_obj.value < 0:
                                    cell_obj.font = fin_red_font

            colorize_sheet(writer.sheets['Detaylı İşlem Listesi'])
            if not df_sheet2.empty:
                colorize_sheet(writer.sheets['Teknik Veriler'])

        return filepath

    def plot_stats_by_feature(self, feature_name: str, bins: int = 8) -> Optional[io.BytesIO]:
        """
        Görsel rapor (Histogram) oluşturur.
        """
        try:
            import matplotlib.pyplot as plt
            import seaborn as sns
        except ImportError:
            return None

        df = self.filter_period('all')
        if df.empty: return None

        data = []
        for _, row in df.iterrows():
            meta = row.get('meta', {}) or {}
            val = None

            if feature_name == 'ai_confidence':
                val = meta.get('ai_confidence')
            elif feature_name == 'technical_score':
                val = meta.get('technical_score')
            elif feature_name == 'momentum':
                val = meta.get('momentum')

            pnl = row.get('realized_net_pct')

            if val is not None and pnl is not None:
                data.append({'val':float(val), 'pnl':float(pnl)})

        if not data: return None

        plot_df = pd.DataFrame(data)

        plt.figure(figsize=(10, 6))
        sns.set_style("whitegrid")

        plot_df['bin'] = pd.cut(plot_df['val'], bins=bins)
        summary = plot_df.groupby('bin', observed=False)['pnl'].mean().reset_index()

        sns.barplot(x='bin', y='pnl', data=summary, palette='RdYlGn')
        plt.title(f"Ortalama PnL vs {feature_name}")
        plt.xticks(rotation=45)
        plt.ylabel("Ortalama PnL (%)")
        plt.tight_layout()

        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        buf.seek(0)
        plt.close()

        return buf
